"""
fetch_eu_prices.py
==================

Downloads the EU's "Weekly Oil Bulletin" prices-with-taxes XLSX,
parses it, and writes data/eu-prices.json for the frontend to fetch.

Run from the project root:

    python3 scripts/fetch_eu_prices.py

Requirements:

    pip install openpyxl

The bulletin updates every Thursday. Re-run this script weekly to refresh.
Later, this same script can live on a server with a cron job.

Source: https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en
The page links to two XLSX files — one with taxes, one without. We use
the WITH-taxes file because that's the price consumers actually pay.

The bulletin reports prices in EUR per 1000 litres. We divide by 1000
so the JSON stores EUR per litre, which is what people are used to seeing
on European pump signs (e.g. "1.85 €/L").
"""

from __future__ import annotations

import json
import sys
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is not installed. Run: pip install openpyxl")

# The European Commission hosts the bulletin behind these long opaque URLs.
# If this 404s in the future, check the bulletin page (in the module docstring)
# for the current "prices with taxes" download link and update XLSX_URL.
XLSX_URL = (
    "https://energy.ec.europa.eu/document/download/"
    "264c2d0f-f161-4ea3-a777-78faae59bea0_en"
    "?filename=Weekly_Oil_Bulletin_Weekly_prices_with_Taxes.xlsx"
)

# Map the country names the bulletin uses to ISO-3166 alpha-2 codes.
# These codes are what data/eu-countries.geojson tags each polygon with,
# so the frontend can match a clicked country to its price by code.
NAME_TO_ISO2 = {
    "Austria": "AT", "Belgium": "BE", "Bulgaria": "BG", "Croatia": "HR",
    "Cyprus": "CY", "Czechia": "CZ", "Denmark": "DK", "Estonia": "EE",
    "Finland": "FI", "France": "FR", "Germany": "DE", "Greece": "GR",
    "Hungary": "HU", "Ireland": "IE", "Italy": "IT", "Latvia": "LV",
    "Lithuania": "LT", "Luxembourg": "LU", "Malta": "MT", "Netherlands": "NL",
    "Poland": "PL", "Portugal": "PT", "Romania": "RO", "Slovakia": "SK",
    "Slovenia": "SI", "Spain": "ES", "Sweden": "SE",
}

# Bulletin column layout (1-indexed):
#   1: country name        2: Euro-super 95 (gasoline)
#   3: automotive gas oil (diesel)
#   4: heating gas oil     5-6: heavy fuel oil   7: LPG
COL_COUNTRY = 1
COL_GASOLINE = 2
COL_DIESEL = 3

# The week-of date sits in row 2, column 1 of the sheet.
ROW_WEEK_DATE = 2
ROW_DATA_START = 3

PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_PATH = PROJECT_ROOT / "data" / "eu-prices.json"


def download(url: str, dest: Path) -> None:
    print(f"downloading {url}")
    req = urllib.request.Request(url, headers={"User-Agent": "gas-prices/0.1"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        dest.write_bytes(resp.read())
    print(f"saved {dest} ({dest.stat().st_size:,} bytes)")


def per_1000l_to_per_l(value):
    if value is None:
        return None
    return round(float(value) / 1000, 3)


def parse(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    week_cell = ws.cell(row=ROW_WEEK_DATE, column=COL_COUNTRY).value
    if isinstance(week_cell, datetime):
        week_of = week_cell.date().isoformat()
    else:
        week_of = str(week_cell) if week_cell else None

    countries: dict[str, dict] = {}
    for r in range(ROW_DATA_START, ws.max_row + 1):
        name = ws.cell(row=r, column=COL_COUNTRY).value
        if not isinstance(name, str):
            continue
        iso2 = NAME_TO_ISO2.get(name.strip())
        if not iso2:
            # The bulletin appends "EUR27 weighted average" and "Euro Area 20
            # weighted average" rows at the bottom. We ignore them — the
            # frontend only cares about individual countries.
            continue
        countries[iso2] = {
            "name": name.strip(),
            "gasoline_eur_per_l": per_1000l_to_per_l(ws.cell(row=r, column=COL_GASOLINE).value),
            "diesel_eur_per_l": per_1000l_to_per_l(ws.cell(row=r, column=COL_DIESEL).value),
        }

    return {
        "updated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "week_of": week_of,
        "source": "EU Weekly Oil Bulletin (prices with taxes)",
        "source_url": "https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en",
        "currency": "EUR",
        "unit": "per_litre",
        "countries": countries,
    }


def main() -> int:
    tmp_xlsx = PROJECT_ROOT / "data" / "_bulletin.xlsx"
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    try:
        download(XLSX_URL, tmp_xlsx)
        payload = parse(tmp_xlsx)
    finally:
        if tmp_xlsx.exists():
            tmp_xlsx.unlink()

    missing = sorted(set(NAME_TO_ISO2.values()) - set(payload["countries"]))
    if missing:
        print(f"warning: no row found for {missing}")

    OUTPUT_PATH.write_text(json.dumps(payload, indent=2))
    print(f"wrote {OUTPUT_PATH} ({len(payload['countries'])} countries, week of {payload['week_of']})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
